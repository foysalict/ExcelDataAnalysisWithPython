import math
from operator import le
from tkinter import Label
from turtle import goto
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import traceback

def PrepareInput(file_path):
    data = pd.read_excel(file_path, skiprows=1, header=None)
    data = data.iloc[:,0:16]
    return data

def generateOutput(output,output_file):
    try:
        print("Writing to file...")
        book = load_workbook(output_file)
        writer = pd.ExcelWriter(output_file, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df1 = pd.DataFrame(output)
        df1.to_excel(writer,"Results" ,index=False,header=False,startrow=1, startcol=0)
        writer.save()
        print("File written.")
        print("Done.")
    except:
        print("Error generated, please check error.txt for more information...")
        with open('error.txt', 'w') as fh:
            traceback.print_exc(file=fh)
    
def samplingInputData(file_path,percentage, clusters):
    try:
        percentage=int(percentage)
        clusters=int(clusters)
        print("Preparing input data for sampling...")
        output = pd.DataFrame()
        data = pd.read_excel(file_path, skiprows=1, header=None)
        data = data.iloc[:,0:16]
        
        total_rows = (data.shape)[0] + 1
        amount_of_sets = round_off_number(percentage / 100 * total_rows, "amount_of_sets")
        amount_for_each_cluster =round_off_number( amount_of_sets / clusters, "amount_for_each_cluster" ) - 1 # amount of sets in each cluster
        
        no_of_even_distance = clusters - 1  
        even_distances = round_off_number((total_rows - amount_of_sets) / no_of_even_distance, "even_distances")
        
        start = 0
        end = amount_for_each_cluster + 1
        for i in range(no_of_even_distance+1):
            output= output.append(data[start:end], ignore_index = True)
            print("Cluster ", i+1, " of", clusters, " done.")
            start = end + even_distances 
            end = start + amount_for_each_cluster + 1
        return output
    except:
        print("Error generated, please check error.txt for more information...")
        with open('error.txt', 'w') as fh:
            traceback.print_exc(file=fh)

def firstPassageTd(data,passage1,passage2,dataLen,inputSeqArray,inputSeqArray02,output_file_path):
    try:
        output = pd.DataFrame()
        data = pd.DataFrame(data)
        passage1Len = len(passage1)
        row = len(data)
        for x in range(passage1Len):
            inputSeqArray.remove(passage1[x])
        i = 0
        j = 0        
        for i in range(row):
            l = 0
            for k in range(passage1Len):
                for j in range(dataLen):
                    if(str(data.values[i][j]) == str(passage1[k])): 
                        l=l+1
                    if(l== passage1Len and j+1==dataLen):
                        for y in range(len(inputSeqArray)):
                            for z in range(dataLen):
                                if(str(data.values[i][z]) == str(inputSeqArray[y])):
                                    output= output.append(data[i:i+1], ignore_index = True)                                
                                    break
                            break  
                    if(k+1 == passage1Len and l != passage1Len and j+1==dataLen):
                         output= output.append(data[i:i+1], ignore_index = True)
            if (i+1 == row):
                break
        if(len(passage2) > 0):
            return secondPassageTd(output,passage2,dataLen,inputSeqArray02,output_file_path)
        else:
            return generateOutput(output,output_file_path)
    except:
        print("Error generated, please check error.txt for more information...")
        with open('error.txt', 'w') as fh:
            traceback.print_exc(file=fh)

def secondPassageTd(data,passage2,dataLen,inputSeqArray02,output_file_path):
    try:
        output = pd.DataFrame()
        data = pd.DataFrame(data)
        passage2Len = len(passage2)
        for t in range(passage2Len):
            inputSeqArray02.remove(passage2[t])
        i = 0
        j = 0
        for i in range(len(data)):
            l = 0
            for k in range(passage2Len):
                for j in range(dataLen):
                    if(str(data.values[i][j]) == str(passage2[k])): 
                        l=l+1
                    if(l== passage2Len and j+1==dataLen):
                        for y in range(len(inputSeqArray02)):
                            for z in range(dataLen):
                                if(str(data.values[i][z]) == str(inputSeqArray02[y])):
                                    output= output.append(data[i:i+1], ignore_index = True)                                
                                    break
                            break
                    if(k+1 == passage2Len and l != passage2Len and j+1==dataLen):
                         output= output.append(data[i:i+1], ignore_index = True)  
            if (i+1 == len(data)):
                break
        return generateOutput(output,output_file_path)
    except:
        print("Error generated, please check error.txt for more information...")
        with open('error.txt', 'w') as fh:
            traceback.print_exc(file=fh)

def round_off_number(num, num_event):
    rounded_num = 0
    if(isinstance( num, int)):
        rounded_num=  num
    int_num = math.trunc(num)
    num_after_decimal = num - int_num
    num_after_decimal_in_string = str(num_after_decimal)[2:4]
    two_num_after_decimal = int(num_after_decimal_in_string)
    if(two_num_after_decimal >= 51):
        rounded_num =  int_num + 1
    else:   
        rounded_num =  int_num
    print(num_event, ": Running number: ", num, " rounded off to: ", rounded_num)
    return rounded_num

def delete_columns(file_path):
    print("Preparing sheet 1...")
    output = pd.DataFrame()
    data = pd.read_excel(file_path, skiprows=1, header=None, usecols=range(0,16))
    output = data.copy()
    print(output.replace(np.nan,""))
    row_no = 1
    for row in output:
        print("Row: ", row_no)
        for i in range(len(output[row])):
            print("deleting value: ", i )
            output[row][i] = ""
        row_no += 1 
    return output


def SamplingProcess(input_file1, template_file ,percentage, clusters):
    try:
        percentage=int(percentage)
        clusters=int(clusters)
        df1 = samplingInputData(input_file1,percentage, clusters)
        print("Writing to file...")
        book = load_workbook(template_file)
        writer = pd.ExcelWriter(template_file, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
       
        df1.to_excel(writer,"Results" ,index=False,header=False,startrow=1, startcol=0)

        writer.save()
        print("File written.")
        print("Done.")
    except:
        print("Error generated, please check error.txt for more information...")
        with open('error.txt', 'w') as fh:
            traceback.print_exc(file=fh)

def delete_main(template_file):
        df1 = delete_columns(template_file)
        print("Deleting columns...")
        book = load_workbook(template_file)
        writer = pd.ExcelWriter(template_file, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
       
        df1.to_excel(writer,"Results" ,index=False,header=False,startrow=1, startcol=0)
        writer.save()
        print("Deleted.")
        print("Done.")


# delete_main( "./out.xlsx")
# main(input_file1="./GUI with results-1.xlsx", template_file = "./out.xlsx" ,percentage=28, clusters=4)

